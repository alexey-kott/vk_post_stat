import dataclasses
import json
import time
# from asyncio import sleep
from pathlib import Path
from time import sleep
import hashlib
from dataclasses import dataclass
from typing import List, Dict, Any, Tuple, Union

from zipfile import ZipFile, ZIP_DEFLATED
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from aiohttp import ClientSession
from channels.generic.websocket import AsyncWebsocketConsumer
from selenium.webdriver import Chrome, ChromeOptions

from vk_post_stat.settings import VK_SERVICE_TOKEN

STATIC_DIR = Path('./static/')


def get_params():
    return {
        'v': '5.84',
        'access_token': VK_SERVICE_TOKEN,
    }


def delay(func):
    sleep(0.34)  # VK.com has limit no more 3 requests per second

    return func


@dataclass
class Post:
    post_link: str  # V
    club_name: str  # V
    club_link: str  # V
    screenshot_name: str
    members_amount: int  # V
    views: int  # V
    likes: int  # V
    # reposts: int  # ВК выключил
    comments_amount: int  # V
    comments: List[Dict[str, Any]]  # V


@delay
async def get_user_info(user_id):
    params = {
        'v': '5.84',
        'access_token': VK_SERVICE_TOKEN,
        'offset': 0,
        'count': 1000
    }
    url = "https://api.vk.com/method/wall.getComments"

    async with ClientSession() as session:
        async with session.get(url, params=params) as response:
            text = await response.text()


@delay
async def get_groups_info(links: List[str]) -> List[Dict[str, Any]]:  # get id, name, screen_name
    group_ids = [link.split('-')[1].split('_')[0] for link in links]
    params = get_params()
    params['group_ids'] = ','.join(group_ids)
    url = "https://api.vk.com/method/groups.getById"

    async with ClientSession() as session:
        async with session.get(url, params=params) as response:
            text = await response.text()
            data = json.loads(text)
            params['group_ids'] = ''

            return data['response']


@delay
async def get_post_comments(owner_id: int, post_id: int) -> Tuple[int, List[Dict[str, Any]]]:
    params = get_params()
    params['owner_id'] = owner_id
    params['post_id'] = post_id
    params['count'] = 1000
    url = "https://api.vk.com/method/wall.getComments"

    async with ClientSession() as session:
        async with session.get(url, params=params) as response:
            text = await response.text()
            data = json.loads(text)

            return data['response']['count'], data['response']['items']


@delay
async def get_post_likes(owner_id: int, item_id: int) -> Tuple[int, List[int]]:
    params = get_params()
    params['type'] = 'post'
    params['owner_id'] = owner_id
    params['item_id'] = item_id
    params['count'] = 1000
    url = "https://api.vk.com/method/likes.getList"

    async with ClientSession() as session:
        async with session.get(url, params=params) as response:
            text = await response.text()
            data = json.loads(text)

            return data['response']['count'], data['response']['items']


@delay
async def get_post_reposts(owner_id: int, post_id: int) -> Tuple[List, List, List]:
    # Сейчас ВКонтакте возвращает пустые поля ответа, вряд ли это когда-то изменится
    # Связана такая ситуация с участившимися случаями посадки за репосты (ДА, ЭТО ПИЗДЕЦ)
    params = get_params()
    params['owner_id'] = owner_id
    params['post_id'] = post_id
    params['count'] = 1000
    url = "https://api.vk.com/method/wall.getReposts"

    async with ClientSession() as session:
        async with session.get(url, params=params) as response:
            text = await response.text()
            data = json.loads(text)

            return data['response']['items'], data['response']['profiles'], data['response']['groups']


@delay
async def get_community_members(group_id: Union[int, str]) -> Tuple[int, List[int]]:
    params = get_params()
    params['group_id'] = str(group_id).strip('-')
    params['count'] = 1000
    url = "https://api.vk.com/method/groups.getMembers"

    async with ClientSession() as session:
        async with session.get(url, params=params) as response:
            text = await response.text()
            data = json.loads(text)

            return data['response']['count'], data['response']['items']


@delay
async def get_post_views(owner_id: int, post_id: int) -> int:
    params = get_params()
    params['posts'] = f"{owner_id}_{post_id}"
    params['count'] = 1000
    url = "https://api.vk.com/method/wall.getById"

    async with ClientSession() as session:
        async with session.get(url, params=params) as response:
            text = await response.text()
            data = json.loads(text)

            return int(data['response'][0]['views']['count'])


@delay
async def get_users_by_id(user_ids: List[int]):
    params = get_params()
    params['user_ids'] = str(user_ids)
    url = "https://api.vk.com/method/users.get"

    async with ClientSession() as session:
        async with session.get(url, params=params) as response:
            text = await response.text()
            data = json.loads(text)

            if data.get('error'):
                return []

            params['user_ids'] = ''
            return data['response']


def write(data: str) -> None:
    with open("data.json", "w") as file:
        file.write(data)


async def send_object_to_client(consumer: AsyncWebsocketConsumer, obj: Any) -> None:
    await consumer.send(text_data=json.dumps({
        'method': 'new_post_stat',
        'value': dataclasses.asdict(obj)
    }))


def get_file_name():
    str_time = str(time.time()).encode('utf-8')
    m = hashlib.md5()
    m.update(str_time)
    name_hash_part = m.hexdigest()[:10]

    return f"vk_post_stat-{name_hash_part}.xlsx"


def generate_excel(posts: List[Post]):
    wb = Workbook()
    ws = wb.create_sheet("Default", index=0)
    ws.append(['Ссылка на запись', 'Название сообщества',
               'Ссылка на сообщество', 'Количество подписчиков',
               'Охват записи', 'Количество лайков', 'Количество комментариев'])

    for i, post in enumerate(posts):
        ws.append([post.post_link, post.club_name, post.club_link, post.members_amount, post.views, post.likes,
                   post.comments_amount])

    for i, column in enumerate(ws.columns):
        ws.column_dimensions[get_column_letter(i + 1)].width = 25

    file_name = Path(get_file_name())
    wb.save(f"./static/reports/{file_name.name}")

    return file_name


def take_screenshot(link: str):
    WINDOW_SIZE = "1220,1200"
    options = ChromeOptions()
    options.add_argument("--start-maximized")
    options.add_argument("--headless")
    options.add_argument("--window-size=%s" % WINDOW_SIZE)
    options.add_argument('--disable-gpu')
    options.add_argument('--disable-dev-shm-usage')
    options.add_argument("--no-sandbox")
    driver = Chrome("./webdriver/chromedriver", chrome_options=options)

    driver.get(link)
    screenshot_name = f"{link.split('-')[1]}.png"
    driver.save_screenshot(f'./static/screenshots/{screenshot_name}')
    driver.close()

    return screenshot_name


def form_archive(screenshot_names, report_name):
    report_path = STATIC_DIR / 'reports' / report_name
    archive_path = STATIC_DIR / 'archives' / (report_name.stem + '.zip')
    with ZipFile(archive_path, 'w', ZIP_DEFLATED) as archive:
        archive.write(report_path)
        for screenshot_name in screenshot_names:
            screenshot_path = STATIC_DIR / 'screenshots' / screenshot_name
            archive.write(screenshot_path)
    archive.close()

    return archive_path.name


async def parse_posts_info(links: List[str], consumer: AsyncWebsocketConsumer) -> str:
    posts = []
    screenshots = []

    groups = await get_groups_info(links)
    id_group_map = {-group['id']: group for group in groups}

    for link in links:
        try:
            owner_id = int(link.split('wall')[1].split('_')[0])  # owner -- group, public page or user
            post_id = int(link.split('-')[1].split('_')[1])

            likes, like_users = await get_post_likes(owner_id, post_id)
            comments_amount, comments = await get_post_comments(owner_id, post_id)
            # repost_items, repost_profiles, repost_groups = await get_post_reposts(owner_id, post_id)
            members, member_items = await get_community_members(owner_id)
            views = await get_post_views(owner_id, post_id)

            screenshots.append(take_screenshot(link))

            post = Post(club_link=f"https://vk.com/{id_group_map[owner_id]['screen_name']}",
                        club_name=id_group_map[owner_id]['name'],
                        screenshot_name=f"{owner_id}_{post_id}.png".strip('-'),
                        comments_amount=comments_amount,
                        members_amount=members,
                        comments=comments,
                        post_link=link,
                        likes=likes,
                        views=views,
                        )
            posts.append(post)

            await send_object_to_client(consumer, post)
        except Exception as e:
            print(e)

    report_name = generate_excel(posts)
    package_name = form_archive(screenshots, report_name)

    return package_name


def save_users(data: Dict[str, List], file_name: str) -> str:
    wb = Workbook()

    for link, users in data.items():
        sheet_name = link.split('-')[1]
        ws = wb.create_sheet(sheet_name)
        ws.append(["Ссылка на пост", link])
        ws.append(["User ID", "Name", "Username"])
        for user in users:
            ws.append([f"https://vk.com/id{user['id']}", user['first_name'], user['last_name']])

    # wb.remove(wb.get_sheet_by_name('Sheet'))
    wb.save(f"./static/reports/{file_name}")
    wb.close()

    return file_name


async def get_like_users(links: List[str]) -> str:
    data_to_save = {}

    for link in links:
        try:
            owner_id = int(link.split('wall')[1].split('_')[0])  # owner -- group, public page or user
            post_id = int(link.split('-')[1].split('_')[1])
            likes, like_users = await get_post_likes(owner_id, post_id)

            users = await get_users_by_id(like_users)

            data_to_save[link] = users
        except Exception as e:
            print(e)

    like_users_file_name = save_users(data_to_save, "like_users.xlsx")

    return like_users_file_name


async def get_comment_users(links: List[str]) -> str:
    data_to_save = {}

    for link in links:
        try:
            owner_id = int(link.split('wall')[1].split('_')[0])  # owner -- group, public page or user
            post_id = int(link.split('-')[1].split('_')[1])

            comments_amount, comments = await get_post_comments(owner_id, post_id)

            users = await get_users_by_id({item['from_id'] for item in comments})

            data_to_save[link] = users
        except Exception as e:
            raise e

    comment_users_file_name = save_users(data_to_save, "comment_users.xlsx")

    return comment_users_file_name
